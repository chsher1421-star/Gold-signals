"""
CME Open Interest auto-updater.

Replicates the workbook's original VBA macro (`download_data` /
`get_all_up_to_date`) in Python so it can run unattended inside a GitHub
Actions workflow, with no Excel / no manual date entry required.

What it does, every run:
  1. Looks at the "Gold" sheet's latest stored date to figure out which
     dates are missing (self-healing catch-up — works even if a run was
     skipped or the file wasn't touched for days).
  2. For each missing date, pulls the CME VOI export (Metals, FX, Energy)
     — tries the Final report first, falls back to Preliminary, exactly
     like the original macro.
  3. Writes the raw data into Metals / FX / Energy sheets, then updates
     every per-instrument sheet + its "* OI Chart" sheet.
  4. Saves the .xlsm in place (macros preserved) so it opens in Excel
     exactly as before.

Config (workbook name, tracked instruments, lookback window) is in
CONFIG below — matches the workbook's own "Links" sheet values.
"""

import io
import sys
import logging
from datetime import date, timedelta

import requests
import pandas as pd
import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
log = logging.getLogger("oi_update")

# ---------------------------------------------------------------------------
# CONFIG — mirrors the workbook's own "Links" sheet
# ---------------------------------------------------------------------------

WORKBOOK_PATH = "oi_automation/OI_Automatic_Update.xlsm"

BASE_URL = "http://www.cmegroup.com/CmeWS/exp/voiProductsViewExport.ctl?media=xls&tradeDate="
ASSET_CLASS = {"metals": "8", "fx": "3", "energy": "7"}
EXCLUDED = "excluded=CEE,CEU,KCB"

DAYS_BACK_MAX = 20  # same safety cap as Links!A15 in the workbook

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.cmegroup.com/market-data/volume-open-interest.html",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "same-origin",
    "Sec-Fetch-Dest": "document",
    "Upgrade-Insecure-Requests": "1",
}

_SESSION = requests.Session()
_SESSION.headers.update(HEADERS)
_WARMED_UP = False


def _warm_up():
    """Visit the CME homepage first to pick up session cookies, like a
    real browser would before hitting the export endpoint directly."""
    global _WARMED_UP
    if _WARMED_UP:
        return
    try:
        _SESSION.get("https://www.cmegroup.com/", timeout=30)
        _SESSION.get(
            "https://www.cmegroup.com/market-data/volume-open-interest.html",
            timeout=30,
        )
    except requests.RequestException as e:
        log.warning(f"Warm-up request failed (continuing anyway): {e}")
    _WARMED_UP = True

# sheet_name -> (source_tab, "Like" pattern translated to python)
# pattern check = startswith(prefix) and (contains substrings in order)
METALS_INSTRUMENTS = {
    "Gold": ("Gold Futures", "starts"),
    "Silver": ("Silver", "Future", "starts_contains"),
    "Copper": ("Copper Future", "starts"),
    "Iron Ore": ("Iron Ore", "(TSI) Future", "starts_contains"),
}
ENERGY_INSTRUMENTS = {
    "Oil": ("Crude Oil Futures", "exact"),
}
FX_INSTRUMENTS = {
    "AUD": ("Australian Dollar Future", "starts"),
    "GBP": ("British Pound Future", "starts"),
    "JPY": ("Japanese Yen Future", "starts"),
    "CHF": ("Swiss Franc Future", "starts"),
    "NZD": ("New Zealand Dollar Future", "starts"),
    "CAD": ("Canadian Dollar Future", "starts"),
    "EUR": ("Euro FX Future", "starts"),
    "EURGBP": ("Euro/British Pound Future", "starts"),
    "EURJPY": ("Euro/Japanese Yen Future", "starts"),
    "EURCHF": ("Euro/Swiss Franc Future", "starts"),
    "EURAUD": ("Euro/Australian Dollar Future", "starts"),
    "EURCAD": ("Euro/Canadian Dollar Future", "starts"),
    "GBPJPY": ("British Pound/Japanese Yen Future", "starts"),
    "AUDJPY": ("Australian Dollar/Japanese Yen Future", "starts"),
    "AUDNZD": ("Australian Dollar/New Zealand Dollar Future", "starts"),
}
CHART_SHEET = {name: f"{name} OI Chart" for name in
               list(METALS_INSTRUMENTS) + list(ENERGY_INSTRUMENTS) + list(FX_INSTRUMENTS)}
CHART_SHEET["EUR"] = "EUR Chart"
for _n in FX_INSTRUMENTS:
    if _n != "EUR":
        CHART_SHEET[_n] = f"{_n} Chart"


def matches(name, rule):
    if not isinstance(name, str):
        return False
    if rule[-1] == "exact":
        return name.strip() == rule[0]
    if rule[-1] == "starts":
        return name.startswith(rule[0])
    if rule[-1] == "starts_contains":
        return name.startswith(rule[0]) and rule[1] in name
    return False


# ---------------------------------------------------------------------------
# Fetching
# ---------------------------------------------------------------------------

def build_url(trade_date: date, asset_class_id: str, report_type: str) -> str:
    d = trade_date.strftime("%Y%m%d")
    return f"{BASE_URL}{d}&assetClassId={asset_class_id}&reportType={report_type}&{EXCLUDED}"


def fetch_table(url: str):
    """Download a CME VOI export and return it as a DataFrame with a
    normalized header row (Name/Type/Globex/Open OutCry/Clear Port/
    Volume/Open Interest/Change), or None if there's no usable data."""
    _warm_up()
    try:
        resp = _SESSION.get(url, timeout=30)
        resp.raise_for_status()
    except requests.RequestException as e:
        log.warning(f"Fetch failed for {url}: {e}")
        return None

    content = resp.content
    if not content or len(content) < 200:
        return None

    df = None
    for attempt in ("excel", "html"):
        try:
            if attempt == "excel":
                df = pd.read_excel(io.BytesIO(content), header=None)
            else:
                tables = pd.read_html(io.BytesIO(content))
                df = tables[0] if tables else None
            if df is not None and not df.empty:
                break
        except Exception:
            df = None
            continue

    if df is None or df.empty:
        return None

    # find the real header row (the one starting with "Name","Type")
    header_row = None
    for i in range(min(15, len(df))):
        row_vals = [str(v).strip() for v in df.iloc[i].tolist()]
        if row_vals[:2] == ["Name", "Type"]:
            header_row = i
            break
    if header_row is None:
        return None

    data = df.iloc[header_row + 1:].reset_index(drop=True)
    data.columns = [
        "Name", "Type", "Globex", "Open OutCry", "Clear Port",
        "Volume", "Open Interest", "Change"
    ][: data.shape[1]]
    data = data.dropna(subset=["Name"])
    for col in ("Volume", "Open Interest", "Change"):
        if col in data.columns:
            data[col] = (
                data[col].astype(str).str.replace(",", "", regex=False)
                .replace("nan", "0")
            )
            data[col] = pd.to_numeric(data[col], errors="coerce").fillna(0).astype(int)
    return data


def fetch_day(trade_date: date):
    """Returns dict {'metals': df|None, 'fx': df|None, 'energy': df|None}
    trying Final report first, then Preliminary, per asset class."""
    out = {}
    for key, class_id in ASSET_CLASS.items():
        df = fetch_table(build_url(trade_date, class_id, "F"))
        if df is None or df.empty:
            df = fetch_table(build_url(trade_date, class_id, "P"))
        out[key] = df
    return out


# ---------------------------------------------------------------------------
# Workbook update
# ---------------------------------------------------------------------------

def find_row(df, rule):
    if df is None:
        return None
    for _, row in df.iterrows():
        if matches(row["Name"], rule):
            return row
    return None


def update_instrument(wb, sheet_name, chart_sheet_name, row, trade_date):
    if row is None:
        return False
    ws = wb[sheet_name]
    last_date = ws.cell(row=2, column=1).value
    oi = int(row["Open Interest"])
    change = int(row["Change"])

    if last_date is not None and hasattr(last_date, "date"):
        last_date = last_date.date()

    if last_date == trade_date:
        ws.cell(row=2, column=3).value = oi
        ws.cell(row=2, column=4).value = change
    elif last_date is None or last_date < trade_date:
        ws.insert_rows(2)
        ws.cell(row=2, column=1).value = trade_date
        ws.cell(row=2, column=1).number_format = "dd/mm/yy;@"
        ws.cell(row=2, column=3).value = oi
        ws.cell(row=2, column=4).value = change
    else:
        # older date than what's already there — nothing to do
        return False

    # refresh the chart-source sheet with the latest 11 points, oldest->newest
    if chart_sheet_name in wb.sheetnames:
        chart_ws = wb[chart_sheet_name]
        for i in range(2, 13):
            src_row = 14 - i
            chart_ws.cell(row=i, column=1).value = ws.cell(row=src_row, column=1).value
            chart_ws.cell(row=i, column=2).value = ws.cell(row=src_row, column=3).value
    return True


def write_raw_sheet(wb, sheet_name, df):
    if df is None or df.empty:
        return
    ws = wb[sheet_name]
    # clear old data range (keep it simple: A1:H5000)
    for r in ws.iter_rows(min_row=1, max_row=5000, min_col=1, max_col=8):
        for c in r:
            c.value = None
    ws.cell(row=5, column=1).value = "Name"
    ws.cell(row=5, column=2).value = "Type"
    ws.cell(row=5, column=3).value = "Globex"
    ws.cell(row=5, column=4).value = "Open OutCry"
    ws.cell(row=5, column=5).value = "Clear Port"
    ws.cell(row=5, column=6).value = "Volume"
    ws.cell(row=5, column=7).value = "Open Interest"
    ws.cell(row=5, column=8).value = "Change"
    for i, (_, row) in enumerate(df.iterrows()):
        r = 6 + i
        ws.cell(row=r, column=1).value = row.get("Name")
        ws.cell(row=r, column=2).value = row.get("Type")
        ws.cell(row=r, column=3).value = row.get("Globex")
        ws.cell(row=r, column=4).value = row.get("Open OutCry")
        ws.cell(row=r, column=5).value = row.get("Clear Port")
        ws.cell(row=r, column=6).value = int(row.get("Volume", 0))
        ws.cell(row=r, column=7).value = int(row.get("Open Interest", 0))
        ws.cell(row=r, column=8).value = int(row.get("Change", 0))


def process_date(wb, trade_date: date) -> bool:
    """Returns True if any data was found/applied for this date."""
    day_data = fetch_day(trade_date)
    if all(df is None for df in day_data.values()):
        log.info(f"{trade_date}: no CME data available (weekend/holiday/not yet published) — skipping")
        return False

    write_raw_sheet(wb, "Metals", day_data["metals"])
    write_raw_sheet(wb, "FX", day_data["fx"])
    write_raw_sheet(wb, "Energy", day_data["energy"])

    applied_any = False
    for sheet, rule in METALS_INSTRUMENTS.items():
        row = find_row(day_data["metals"], rule)
        if update_instrument(wb, sheet, CHART_SHEET[sheet], row, trade_date):
            applied_any = True
    for sheet, rule in ENERGY_INSTRUMENTS.items():
        row = find_row(day_data["energy"], rule)
        if update_instrument(wb, sheet, CHART_SHEET[sheet], row, trade_date):
            applied_any = True
    for sheet, rule in FX_INSTRUMENTS.items():
        row = find_row(day_data["fx"], rule)
        if update_instrument(wb, sheet, CHART_SHEET[sheet], row, trade_date):
            applied_any = True

    if applied_any:
        log.info(f"{trade_date}: workbook updated")
    else:
        log.info(f"{trade_date}: data fetched but no tracked instruments matched")
    return applied_any


def main():
    log.info(f"Opening workbook: {WORKBOOK_PATH}")
    wb = openpyxl.load_workbook(WORKBOOK_PATH, keep_vba=True, data_only=False)

    gold_ws = wb["Gold"]
    last_known = gold_ws.cell(row=2, column=1).value
    if hasattr(last_known, "date"):
        last_known = last_known.date()
    today = date.today()

    if last_known is None:
        start = today - timedelta(days=DAYS_BACK_MAX)
    else:
        start = last_known + timedelta(days=1)
        if (today - start).days > DAYS_BACK_MAX:
            start = today - timedelta(days=DAYS_BACK_MAX)

    if start > today:
        log.info("Workbook already up to date — nothing to do.")
        return

    any_update = False
    d = start
    while d <= today:
        if process_date(wb, d):
            any_update = True
        d += timedelta(days=1)

    if any_update:
        wb.save(WORKBOOK_PATH)
        log.info("Saved workbook.")
    else:
        log.info("No new data found for any missing date — workbook left unchanged.")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        log.exception("Update failed")
        sys.exit(1)
